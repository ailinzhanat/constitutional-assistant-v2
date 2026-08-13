"""
Модуль общей аналитики и анализа обратной связи — Constitutional Assistant.

Реализует ТЗ "Модуль общей аналитики и анализа обратной связи":
- сбор обезличенных событий (посещаемость, воронка шагов 01-04)
- агрегация метрик для дашборда
- категоризация и тональность отзывов (через Groq, опционально)
- экспорт отчёта в PDF/Excel

Хранение: Neo4j (та же база, что уже используется в проекте для законов/статей).
Выбрано вместо JSONL-файлов, потому что Render free-tier имеет эфемерную
файловую систему — файлы стираются при каждом рестарте/переразвёртывании
сервиса. Neo4j Aura (даже на бесплатном плане) данные не теряет: инстанс
может "засыпать" при бездействии, но при пробуждении (Resume в консоли
Neo4j Aura) все ранее записанные данные остаются на месте.

Как подключить в main.py — см. комментарии внизу файла ("INTEGRATION").
"""

import json
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from typing import Callable, Literal, Optional

from fastapi import APIRouter, Header, HTTPException
from pydantic import BaseModel, Field

import os

ADMIN_ACCESS_CODE = os.environ.get("ADMIN_ACCESS_CODE", "")  # тот же код, что и в feedback.py

router = APIRouter(prefix="/api/analytics", tags=["analytics"])

# Заполняется через init_analytics_router(run_query_fn) при подключении
# роутера в main.py — так модуль переиспользует уже существующее
# подключение к Neo4j вместо того чтобы открывать своё собственное.
_run_query: Optional[Callable] = None


def init_analytics_router(run_query_fn: Callable):
    """
    Вызывается один раз из main.py при старте приложения.
    run_query_fn(query: str, params: dict) -> List[Dict] — та же функция
    run_query, что уже используется в main.py для запросов к Neo4j.
    """
    global _run_query
    _run_query = run_query_fn


def _query(cypher: str, params: dict = None):
    if _run_query is None:
        raise HTTPException(
            status_code=503,
            detail="Analytics router не инициализирован (init_analytics_router не вызван в main.py)"
        )
    try:
        return _run_query(cypher, params or {})
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Ошибка подключения к Neo4j: {str(e)}")


def _check_admin(x_admin_code: Optional[str]):
    if not ADMIN_ACCESS_CODE or x_admin_code != ADMIN_ACCESS_CODE:
        raise HTTPException(status_code=403, detail="Неверный код доступа администратора")


# ---------------------------------------------------------------------------
# Модели
# ---------------------------------------------------------------------------

StepId = Literal["01_consent", "02_description", "03_jurisdiction", "04_draft"]
EventType = Literal["step_started", "step_completed", "step_abandoned", "draft_downloaded"]

# FR-7: причины отказа на шаге 03 (соответствуют подп.1-5 п.2 ст.47 Конституционного закона)
JurisdictionReason = Literal[
    "within_jurisdiction",
    "not_participant",           # не участник дела
    "wrong_object",               # оспаривается акт применения, а не норма
    "deadline_expired",           # истёк годичный срок
    "no_valid_poa",                # нет/неполная доверенность
    "improper_subject",           # ненадлежащий субъект (напр. юрлицо)
    "other_out_of_jurisdiction",
]


class AnalyticsEvent(BaseModel):
    session_id: str = Field(..., description="Случайный анонимный ID сессии (генерируется на фронте, НЕ привязан к личности)")
    event_type: EventType
    step: StepId
    language: Literal["kk", "ru", "en"]
    device: Literal["mobile", "desktop"]
    jurisdiction_reason: Optional[JurisdictionReason] = None  # заполняется только для шага 03
    region: Optional[str] = None  # агрегированный регион, если пользователь дал согласие; без точной геолокации
    is_new_visitor: bool = True
    timestamp: Optional[str] = None  # если не передано — берём серверное время


class SurveyResponse(BaseModel):
    satisfaction: int = Field(..., ge=1, le=5)
    nps: int = Field(..., ge=0, le=10)
    completion_time_seconds: Optional[int] = None
    language: Literal["kk", "ru", "en"]
    timestamp: Optional[str] = None


# ---------------------------------------------------------------------------
# FR-1..FR-4: события посещаемости / воронки (публичный, обезличенный)
# ---------------------------------------------------------------------------

@router.post("/event")
def record_event(event: AnalyticsEvent):
    """
    Публичный эндпоинт — вызывается фронтендом при каждом переходе по шагам.
    NFR-1: никаких ФИО/IP/точной геолокации не принимается и не хранится.
    1. Записывается как узел :AnalyticsEvent в Neo4j (основное хранилище)
    2. Дополнительно пишется в Google Sheets, вкладка "Analytics" (резервная копия)
    """
    record = event.model_dump()
    record["timestamp"] = record["timestamp"] or datetime.now(timezone.utc).isoformat()
    record["id"] = str(uuid.uuid4())

    _query(
        """
        CREATE (e:AnalyticsEvent {
            id: $id,
            session_id: $session_id,
            event_type: $event_type,
            step: $step,
            language: $language,
            device: $device,
            jurisdiction_reason: $jurisdiction_reason,
            region: $region,
            is_new_visitor: $is_new_visitor,
            timestamp: $timestamp
        })
        """,
        record,
    )

    # Резервная копия в Google Sheets — не должна ронять запрос, если
    # основная запись в Neo4j уже прошла успешно.
    try:
        from sheets_writer import write_analytics_event_to_sheets
        sheets_result = write_analytics_event_to_sheets(
            session_id=record["session_id"],
            step=record["step"],
            event_type=record["event_type"],
            language=record["language"],
            device=record["device"],
            jurisdiction_reason=record.get("jurisdiction_reason"),
            is_new_visitor=record["is_new_visitor"],
        )
        if not sheets_result.get("success"):
            print(f"⚠️ Sheets analytics write failed: {sheets_result.get('error')}")
    except Exception as e:
        print(f"⚠️ Sheets analytics import/write error: {e}")

    return {"status": "ok"}


@router.post("/survey")
def record_survey(response: SurveyResponse):
    """FR-13: приём ответов опросника (см. ТЗ «Опросник для респондентов»)."""
    record = response.model_dump()
    record["timestamp"] = record["timestamp"] or datetime.now(timezone.utc).isoformat()
    record["id"] = str(uuid.uuid4())

    _query(
        """
        CREATE (s:AnalyticsSurvey {
            id: $id,
            satisfaction: $satisfaction,
            nps: $nps,
            completion_time_seconds: $completion_time_seconds,
            language: $language,
            timestamp: $timestamp
        })
        """,
        record,
    )
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# Агрегация для дашборда (защищено X-Admin-Code, см. NFR-3)
# ---------------------------------------------------------------------------

STEP_ORDER: list[StepId] = ["01_consent", "02_description", "03_jurisdiction", "04_draft"]


def _since_iso(period_days: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=period_days)).isoformat()


def _fetch_events(since_iso: str) -> list[dict]:
    rows = _query(
        """
        MATCH (e:AnalyticsEvent)
        WHERE e.timestamp >= $since
        RETURN e.id AS id, e.session_id AS session_id, e.event_type AS event_type,
               e.step AS step, e.language AS language, e.device AS device,
               e.jurisdiction_reason AS jurisdiction_reason, e.is_new_visitor AS is_new_visitor,
               e.timestamp AS timestamp
        """,
        {"since": since_iso},
    )
    return rows


def _fetch_surveys(since_iso: str) -> list[dict]:
    rows = _query(
        """
        MATCH (s:AnalyticsSurvey)
        WHERE s.timestamp >= $since
        RETURN s.satisfaction AS satisfaction, s.nps AS nps, s.timestamp AS timestamp
        """,
        {"since": since_iso},
    )
    return rows


@router.get("/dashboard")
def get_dashboard(period_days: int = 30, x_admin_code: Optional[str] = Header(None)):
    """
    FR-14: сводная панель — единый ответ со всеми метриками для дашборда.
    Малая выборка (риск из ТЗ п.6): всегда возвращаем абсолютные числа
    наравне с процентами.
    """
    _check_admin(x_admin_code)

    since_iso = _since_iso(period_days)
    events = _fetch_events(since_iso)
    surveys = _fetch_surveys(since_iso)

    # --- FR-1/FR-2: посещаемость ---
    sessions = {e["session_id"] for e in events}
    new_sessions = {e["session_id"] for e in events if e.get("is_new_visitor")}
    unique_visitors = len(sessions)
    new_visitors = len(new_sessions)
    returning_visitors = unique_visitors - new_visitors

    # --- FR-3: язык ---
    language_breakdown = Counter(e["language"] for e in events)

    # --- FR-4: устройство ---
    device_breakdown = Counter(e["device"] for e in events)

    # --- FR-5/FR-6: воронка ---
    sessions_per_step: dict[str, set] = defaultdict(set)
    completed_per_step: dict[str, set] = defaultdict(set)
    for e in events:
        if e["event_type"] in ("step_started", "step_completed"):
            sessions_per_step[e["step"]].add(e["session_id"])
        if e["event_type"] == "step_completed":
            completed_per_step[e["step"]].add(e["session_id"])

    funnel = []
    for step in STEP_ORDER:
        reached = len(sessions_per_step[step])
        completed = len(completed_per_step[step])
        drop_off = reached - completed
        drop_off_rate = round(drop_off / reached * 100, 1) if reached else 0.0
        funnel.append({
            "step": step,
            "reached": reached,
            "completed": completed,
            "drop_off": drop_off,
            "drop_off_rate_pct": drop_off_rate,
        })

    # --- FR-7: разбивка решений шага 03 ---
    jurisdiction_reasons = Counter(
        e["jurisdiction_reason"] for e in events
        if e["step"] == "03_jurisdiction" and e.get("jurisdiction_reason")
    )

    # --- FR-8: черновики скачаны ---
    drafts_downloaded = sum(1 for e in events if e["event_type"] == "draft_downloaded")

    # --- FR-13: опросник ---
    if surveys:
        avg_satisfaction = round(sum(s["satisfaction"] for s in surveys) / len(surveys), 2)
        promoters = sum(1 for s in surveys if s["nps"] >= 9)
        detractors = sum(1 for s in surveys if s["nps"] <= 6)
        nps_score = round((promoters - detractors) / len(surveys) * 100, 1)
    else:
        avg_satisfaction = None
        nps_score = None

    # --- FR-14: сводка одним взглядом ---
    step04 = next(s for s in funnel if s["step"] == "04_draft")
    step01 = next(s for s in funnel if s["step"] == "01_consent")
    overall_completion_rate = (
        round(step04["completed"] / step01["reached"] * 100, 1) if step01["reached"] else 0.0
    )
    out_of_jurisdiction = sum(v for k, v in jurisdiction_reasons.items() if k != "within_jurisdiction")
    total_jurisdiction_checks = sum(jurisdiction_reasons.values())
    out_of_jurisdiction_rate = (
        round(out_of_jurisdiction / total_jurisdiction_checks * 100, 1) if total_jurisdiction_checks else 0.0
    )

    return {
        "period_days": period_days,
        "summary": {
            "unique_visitors": unique_visitors,
            "overall_completion_rate_pct": overall_completion_rate,
            "out_of_jurisdiction_rate_pct": out_of_jurisdiction_rate,
            "avg_satisfaction": avg_satisfaction,
            "nps_score": nps_score,
        },
        "visitors": {
            "unique": unique_visitors,
            "new": new_visitors,
            "returning": returning_visitors,
            "by_language": dict(language_breakdown),
            "by_device": dict(device_breakdown),
        },
        "funnel": funnel,
        "jurisdiction_breakdown": dict(jurisdiction_reasons),
        "drafts_downloaded": drafts_downloaded,
        "survey": {
            "responses_count": len(surveys),
            "avg_satisfaction": avg_satisfaction,
            "nps_score": nps_score,
        },
    }


# ---------------------------------------------------------------------------
# FR-15: экспорт отчёта PDF/Excel
# ---------------------------------------------------------------------------

@router.get("/export")
def export_report(
    period_days: int = 30,
    fmt: Literal["pdf", "excel"] = "excel",
    x_admin_code: Optional[str] = Header(None),
):
    """
    NFR: экспортированный отчёт не должен содержать персональных данных —
    он строится исключительно из уже агрегированных данных get_dashboard().
    """
    _check_admin(x_admin_code)
    data = get_dashboard(period_days=period_days, x_admin_code=x_admin_code)

    if fmt == "excel":
        return _export_excel(data)
    return _export_pdf(data)


def _export_excel(data: dict):
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Метрика", "Значение"])
    for k, v in data["summary"].items():
        ws.append([k, v])

    ws2 = wb.create_sheet("Воронка")
    ws2.append(["Шаг", "Дошли", "Завершили", "Отвал", "% отвала"])
    for row in data["funnel"]:
        ws2.append([row["step"], row["reached"], row["completed"], row["drop_off"], row["drop_off_rate_pct"]])

    ws3 = wb.create_sheet("Юрисдикция")
    ws3.append(["Причина", "Количество"])
    for k, v in data["jurisdiction_breakdown"].items():
        ws3.append([k, v])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=analytics_report.xlsx"},
    )


def _export_pdf(data: dict):
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 50

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Constitutional Assistant — отчёт аналитики")
    y -= 30

    c.setFont("Helvetica", 11)
    for k, v in data["summary"].items():
        c.drawString(50, y, f"{k}: {v}")
        y -= 18

    y -= 10
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Воронка по шагам")
    y -= 20
    c.setFont("Helvetica", 10)
    for row in data["funnel"]:
        c.drawString(50, y, f"{row['step']}: дошли {row['reached']}, завершили {row['completed']}, отвал {row['drop_off_rate_pct']}%")
        y -= 16

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=analytics_report.pdf"},
    )


# ---------------------------------------------------------------------------
# FR-9..FR-12: обратная связь — категоризация и тональность
# ---------------------------------------------------------------------------

FEEDBACK_CATEGORIES = [
    "непонятная формулировка",
    "ошибка распознавания текста",
    "неверное определение юрисдикции",
    "техническая ошибка",
    "предложение",
    "другое",
]


def categorize_feedback_with_groq(text: str) -> dict:
    """
    FR-10/FR-11: категоризация + тональность через Groq.
    ВАЖНО (риск из ТЗ п.6): используется только как подсказка модератору,
    не как окончательное решение — категория остаётся редактируемой.
    Если Groq недоступен — возвращает 'другое'/'нейтрально' без падения.
    """
    try:
        from groq_analyzer import get_groq_client  # переиспользуем клиент проекта
        client = get_groq_client()

        prompt = f"""Проанализируй отзыв пользователя юридического сервиса.
Категории: {", ".join(FEEDBACK_CATEGORIES)}.
Тональность: позитив / нейтрально / негатив.
Ответь СТРОГО в формате JSON: {{"category": "...", "sentiment": "..."}}

Отзыв: {text}"""

        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=100,
        )
        result = json.loads(response.choices[0].message.content)
        return {
            "category": result.get("category", "другое"),
            "sentiment": result.get("sentiment", "нейтрально"),
            "auto_categorized": True,
        }
    except Exception:
        return {"category": "другое", "sentiment": "нейтрально", "auto_categorized": False}


# ---------------------------------------------------------------------------
# INTEGRATION — как подключить в main.py
# ---------------------------------------------------------------------------
#
# 1. Скопировать этот файл в backend/api/analytics.py (заменить старую версию)
#
# 2. В main.py добавить импорт и подключение роутера — В main.py должно
#    появиться следующее (учитывая, что там уже есть функция run_query):
#
#      from analytics import router as analytics_router, init_analytics_router
#      app.include_router(analytics_router)
#      init_analytics_router(run_query)
#
#    Порядок важен: init_analytics_router(run_query) можно вызвать сразу
#    после определения функции run_query в main.py (она не обязана быть
#    внутри startup — Python передаёт саму функцию, а не её результат).
#
# 3. В requirements.txt добавить (если ещё нет):
#      openpyxl
#      reportlab
#
# 4. Никаких новых .gitignore записей не требуется — данные больше не
#    пишутся в локальные файлы.
#
# 5. На фронтенде (chat.html) события отправляются как раньше — сам
#    формат запроса /api/analytics/event не изменился.
#
# 6. Рекомендуется (не обязательно) создать индексы в Neo4j для скорости:
#      CREATE INDEX IF NOT EXISTS FOR (e:AnalyticsEvent) ON (e.timestamp)
#      CREATE INDEX IF NOT EXISTS FOR (e:AnalyticsEvent) ON (e.session_id)
#    Выполнить один раз через Neo4j Browser / Aura Console → Query tab.
#
# ВАЖНО: у бесплатного Neo4j Aura инстанс "засыпает" при долгом
# бездействии — тогда запросы к дашборду будут падать с ошибкой
# 503 "Ошибка подключения к Neo4j", пока кто-то вручную не нажмёт
# "Resume" в консоли console.neo4j.io. Сами данные при этом НЕ теряются —
# это отличие от прежней JSONL-версии, где рестарт Render стирал файлы.
